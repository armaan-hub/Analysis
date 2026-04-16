"""
Trial Balance Mapper — normalises any client Excel/CSV Trial Balance to the
standard internal schema.

Standard output rows:
  {
    "account_code": str,
    "account_name": str,
    "category":     str,   # assets | liabilities | equity | revenue | expenses
    "debit":        float,
    "credit":       float,
    "net":          float, # debit - credit
  }
"""

from __future__ import annotations

import re
import json
import logging
from typing import Optional

logger = logging.getLogger(__name__)

# ── Keyword maps ──────────────────────────────────────────────────────────────

CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "assets": [
        "asset", "cash", "bank", "receivable", "inventory", "stock",
        "property", "equipment", "machinery", "vehicle", "building",
        "prepaid", "deposit", "investment", "goodwill", "intangible",
    ],
    "liabilities": [
        "liability", "payable", "loan", "overdraft", "provision",
        "deferred", "creditor", "borrowing", "debt", "lease liability",
        "vat payable", "tax payable",
    ],
    "equity": [
        "equity", "capital", "retained", "reserve", "share",
        "partner", "owner", "dividend", "drawing",
    ],
    "revenue": [
        "revenue", "sales", "turnover", "rent income",
        "interest income", "other income", "service income",
        "commission received", "fee received", "fee income",
    ],
    "expenses": [
        "expense", "cost", "depreciation", "amortisation", "amortization",
        "salaries", "wages", "rent", "utilities", "insurance", "repairs",
        "maintenance", "marketing", "advertising", "travel", "legal",
        "professional", "bank charge", "interest expense", "loss",
        "fee", "fees", "commission expense",
    ],
}

COLUMN_ALIASES: dict[str, list[str]] = {
    "account_code": [
        "account code", "acct code", "gl code", "code", "account no",
        "account number", "acct no", "a/c code", "a/c no",
    ],
    "account_name": [
        "account name", "account title", "description", "ledger",
        "account description", "particulars", "narration", "name",
    ],
    "debit": [
        "debit", "dr", "debit amount", "debit balance", "debit total",
    ],
    "credit": [
        "credit", "cr", "credit amount", "credit balance", "credit total",
    ],
}


# ── Tally Group → Financial Category Mapping ──────────────────────────────────
# Maps common Tally group names to IFRS-aligned categories.
# The category determines which financial statement section the account goes to.

TALLY_GROUP_CATEGORY: dict[str, str] = {
    "capital account": "equity",
    "reserves & surplus": "equity",
    "partners capital": "equity",
    "share capital": "equity",
    "loans (liability)": "liabilities",
    "secured loans": "liabilities",
    "unsecured loans": "liabilities",
    "current liabilities": "liabilities",
    "provisions": "liabilities",
    "duties & taxes": "liabilities",
    "fixed assets": "assets_non_current",
    "investments": "assets_non_current",
    "current assets": "assets_current",
    "bank accounts": "assets_current",
    "cash-in-hand": "assets_current",
    "deposits (asset)": "assets_current",
    "sundry debtors": "assets_current",
    "stock-in-hand": "assets_current",
    "loans & advances (asset)": "assets_current",
    "direct incomes": "revenue",
    "direct incomes (income (direct))": "revenue",
    "sales accounts": "revenue",
    "indirect incomes": "other_income",
    "indirect incomes (income (indirect))": "other_income",
    "direct expenses": "cost_of_sales",
    "direct expenses (expenses (direct))": "cost_of_sales",
    "purchase accounts": "cost_of_sales",
    "indirect expenses": "expenses",
    "indirect expenses (expenses (indirect))": "expenses",
    "manufacturing expenses": "cost_of_sales",
}

# Rows whose name exactly matches these are Tally meta-rows to exclude.
_EXCLUDE_ROWS = {"grand total", "total", "nett total", "difference in opening balances"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalise(s: object) -> str:
    """Lower-case, strip, collapse spaces."""
    return re.sub(r"\s+", " ", str(s).lower().strip())


def _match_column(raw: str, aliases: list[str]) -> bool:
    n = _normalise(raw)
    return any(alias in n or n in alias for alias in aliases)


def _read_tally_hierarchy(file_path: str) -> list[dict]:
    """
    Read a Tally-exported Excel trial balance, using bold formatting to detect
    group headers vs ledger (actual) accounts.

    Tally TB structure:
      - Header area (company name, address, date range, column headers)
      - Column A = Particulars, Column B = Debit, Column C = Credit
      - Bold rows = Tally group headers (contain subtotals → skip)
      - Non-bold rows = actual ledger accounts (classify by parent group)

    Returns a list of dicts:
      {"name": str, "is_group": bool, "parent_group": str|None,
       "category": str, "debit": float, "credit": float}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Step 1: Find the header row containing "Debit" / "Credit" to locate
    # the actual column positions and the start of data.
    debit_col: int | None = None
    credit_col: int | None = None
    data_start_row = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=False), 1):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            v_lower = str(v).strip().lower()
            if v_lower == "debit":
                debit_col = cell.column  # 1-indexed column number
            elif v_lower == "credit":
                credit_col = cell.column
        if debit_col and credit_col:
            data_start_row = row_idx + 1
            break

    # Fallback: assume B=Debit, C=Credit if headers not found
    if debit_col is None:
        debit_col = 2
    if credit_col is None:
        credit_col = 3

    logger.info("Tally TB: debit=col%d, credit=col%d, data starts row %d",
                debit_col, credit_col, data_start_row)

    # Step 2: Read data rows from data_start_row onwards
    entries: list[dict] = []
    current_group: str | None = None
    current_group_category: str = "other"

    def _cell_float(cell) -> float:
        """Safely extract float from a cell."""
        v = cell.value
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return 0.0

    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=False):
        cells = list(row)
        # Get account name from column A (index 0)
        name_cell = cells[0] if cells else None
        # Handle merged cells that have no value
        if name_cell is None:
            continue
        name_val = str(name_cell.value).strip() if name_cell.value is not None else ""
        if not name_val or len(name_val) < 2:
            continue

        name_lower = name_val.lower().strip()
        if name_lower in _EXCLUDE_ROWS:
            continue

        # Detect bold (group header)
        is_bold = False
        if hasattr(name_cell, "font") and name_cell.font:
            is_bold = bool(name_cell.font.bold)

        # Read debit and credit from specific columns
        debit = 0.0
        credit = 0.0
        for cell in cells:
            try:
                col_num = cell.column
            except AttributeError:
                continue  # Skip MergedCell objects without column
            if col_num == debit_col:
                debit = _cell_float(cell)
            elif col_num == credit_col:
                credit = _cell_float(cell)

        if is_bold:
            current_group = name_val
            current_group_category = TALLY_GROUP_CATEGORY.get(name_lower, "other")
            entries.append({
                "name": name_val,
                "is_group": True,
                "parent_group": None,
                "category": current_group_category,
                "debit": round(debit, 2),
                "credit": round(credit, 2),
            })
        else:
            entries.append({
                "name": name_val,
                "is_group": False,
                "parent_group": current_group,
                "category": current_group_category,
                "debit": round(debit, 2),
                "credit": round(credit, 2),
            })

    wb.close()
    return entries


def _detect_header_row(df) -> int:
    """
    Return index of the most likely column-header row.

    Strategy (in priority order):
    1. First row that contains both 'debit' and 'credit' keywords → strong signal.
    2. First row that contains at least one of: debit/credit/balance/particulars/account.
    3. First row with ≥3 meaningful (non-empty, non-numeric, non-'nan') text cells.
    4. Fall back to row 0.
    """
    import pandas as pd
    _EMPTY = {"", "nan", "none", "n/a", "null"}
    _STRONG = {"debit", "credit"}
    _WEAK = {"debit", "credit", "balance", "particulars", "account", "description",
             "ledger", "name", "amount", "dr", "cr", "narration", "details"}

    first_weak: int | None = None
    first_triple: int | None = None

    for i, row in df.iterrows():
        row_texts = [
            str(v).strip().lower()
            for v in row
            if v is not None
            and not (isinstance(v, float) and pd.isna(v))
            and str(v).strip().lower() not in _EMPTY
            and not str(v).strip().replace(".", "").replace("-", "").replace(",", "").isdigit()
        ]
        row_set = set(row_texts)

        # Priority 1: row contains BOTH debit and credit
        if _STRONG.issubset(row_set):
            return int(i)

        # Priority 2: row contains at least one financial keyword
        if first_weak is None and row_set & _WEAK:
            first_weak = int(i)

        # Priority 3: row has ≥3 meaningful text cells
        if first_triple is None and len(row_texts) >= 3:
            first_triple = int(i)

    return first_weak if first_weak is not None else (first_triple if first_triple is not None else 0)


def _fuzzy_map_columns(headers: list[str]) -> dict[str, Optional[str]]:
    """Return {standard_field: raw_column_name | None}."""
    mapping: dict[str, Optional[str]] = {k: None for k in COLUMN_ALIASES}
    for raw in headers:
        for std_field, aliases in COLUMN_ALIASES.items():
            if mapping[std_field] is None and _match_column(raw, aliases):
                mapping[std_field] = raw
    return mapping


def _classify_account(name: str) -> str:
    """Classify account using longest-keyword-match-wins strategy."""
    n = _normalise(name)
    best_cat = "other"
    best_len = 0
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in n and len(kw) > best_len:
                best_len = len(kw)
                best_cat = category
    # Special case: "income" alone (e.g. "Direct Incomes") is revenue,
    # but "income" inside expense context stays expense
    if best_cat == "other" and "income" in n:
        best_cat = "revenue"
    if best_cat == "other" and ("gain" in n or "commission" in n):
        # Bare "commission" without "expense" is revenue
        if "expense" not in n and "cost" not in n:
            best_cat = "revenue"
    return best_cat


def _to_float(val: object) -> float:
    try:
        if val is None:
            return 0.0
        import pandas as pd
        if isinstance(val, float) and pd.isna(val):
            return 0.0
        return float(str(val).replace(",", "").strip() or 0)
    except (ValueError, TypeError):
        return 0.0


# ── Public API ────────────────────────────────────────────────────────────────

def map_trial_balance(file_path: str) -> list[dict]:
    """
    Read any .xlsx / .xls / .csv Trial Balance and return normalised rows.

    For Tally-exported Excel files (detected via bold group headers), uses the
    group hierarchy to classify accounts accurately and skips group subtotal rows.

    For other formats, falls back to pandas-based header detection + keyword
    classification.

    Returns:
        List of dicts with keys: account_code, account_name, category,
        debit, credit, net.
    """
    ext = file_path.rsplit(".", 1)[-1].lower()

    # Try Tally-aware reader for Excel files
    if ext in ("xlsx", "xls"):
        try:
            tally_entries = _read_tally_hierarchy(file_path)
            # Check if we actually found bold group headers (Tally format)
            has_groups = any(e["is_group"] for e in tally_entries)
            if has_groups:
                rows = []
                for e in tally_entries:
                    if e["is_group"]:
                        continue  # Skip group subtotal rows
                    name = e["name"]
                    name_lower = name.lower().strip()
                    # Skip meta rows
                    if name_lower in _EXCLUDE_ROWS:
                        continue
                    # Use parent group category; refine with keyword matching if needed
                    category = e["category"]
                    if category == "other":
                        category = _classify_account(name)
                    rows.append({
                        "account_code": "",
                        "account_name": name,
                        "category": category,
                        "debit": e["debit"],
                        "credit": e["credit"],
                        "net": round(e["debit"] - e["credit"], 2),
                    })
                if rows:
                    logger.info(
                        "Trial Balance mapped (Tally hierarchy): %d accounts "
                        "(%d group headers skipped)",
                        len(rows),
                        sum(1 for e in tally_entries if e["is_group"]),
                    )
                    return rows
        except Exception as exc:
            logger.warning("Tally hierarchy detection failed, falling back: %s", exc)

    # Fallback: pandas-based approach for non-Tally or CSV files
    return _map_trial_balance_pandas(file_path)


def _map_trial_balance_pandas(file_path: str) -> list[dict]:
    """Pandas-based TB mapping (original approach for non-Tally files)."""
    import pandas as pd

    ext = file_path.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        df_raw = pd.read_csv(file_path, header=None, dtype=str, keep_default_na=False)
    else:
        df_raw = _best_sheet_df(file_path)

    # Detect the actual header row
    header_row = _detect_header_row(df_raw)

    # Use df_raw with the detected header row — set column names from that row
    # then drop all rows up to and including the header row.
    _EMPTY_NAMES = {"nan", "none", ""}
    raw_col_names = [str(v).strip() for v in df_raw.iloc[header_row].tolist()]
    # Deduplicate: replace "nan"/empty names with unique placeholders
    seen: dict[str, int] = {}
    col_names = []
    for name in raw_col_names:
        if name.lower() in _EMPTY_NAMES:
            key = "__unnamed__"
            seen[key] = seen.get(key, -1) + 1
            col_names.append(f"__col_{seen[key]}__")
        else:
            seen[name] = seen.get(name, -1) + 1
            col_names.append(f"{name}_{seen[name]}" if seen[name] > 0 else name)

    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = col_names
    df = df.reset_index(drop=True)

    if ext == "csv":
        # CSV: re-read with correct header for cleaner dtypes
        df = pd.read_csv(file_path, header=header_row, dtype=str, keep_default_na=False)

    df.columns = [str(c) for c in df.columns]

    col_map = _fuzzy_map_columns(list(df.columns))

    # Fallback: if account_name not mapped, use the first column that looks like text descriptions
    if col_map["account_name"] is None:
        import pandas as pd
        _SKIP = {"nan", "none", ""}
        for col in df.columns:
            text_count = sum(
                1 for v in df[col]
                if v is not None
                and not (isinstance(v, float) and pd.isna(v))
                and str(v).strip().lower() not in _SKIP
                and not str(v).strip().replace(".", "").replace("-", "").replace(",", "").isdigit()
            )
            if text_count >= 3:
                col_map["account_name"] = col
                logger.info(f"Auto-detected account_name column: '{col}' ({text_count} text rows)")
                break

    logger.info(f"TB column mapping: {col_map}")

    rows: list[dict] = []
    for _, row in df.iterrows():
        account_name_raw = row.get(col_map["account_name"] or "", "")
        if not account_name_raw or _normalise(account_name_raw) in ("", "nan", "none"):
            continue

        account_code = str(row.get(col_map["account_code"] or "", "")).strip() if col_map["account_code"] else ""
        account_name = str(account_name_raw).strip()
        debit = _to_float(row.get(col_map["debit"] or "", 0) if col_map["debit"] else 0)
        credit = _to_float(row.get(col_map["credit"] or "", 0) if col_map["credit"] else 0)

        rows.append({
            "account_code": account_code,
            "account_name": account_name,
            "category": _classify_account(account_name),
            "debit": round(debit, 2),
            "credit": round(credit, 2),
            "net": round(debit - credit, 2),
        })

    if not rows:
        raise ValueError("No data rows found in Trial Balance. Check column headers and file format.")

    logger.info(f"Trial Balance mapped: {len(rows)} accounts")
    return rows


def _best_sheet_df(file_path: str):
    """
    For multi-sheet Excel workbooks, return the DataFrame from the sheet that
    most likely contains structured financial data (Debit/Credit or most rows).
    Falls back to the first sheet.
    """
    import pandas as pd
    try:
        xl = pd.ExcelFile(file_path)
    except Exception:
        return pd.read_excel(file_path, header=None, dtype=str)

    if len(xl.sheet_names) == 1:
        return pd.read_excel(file_path, sheet_name=xl.sheet_names[0], header=None, dtype=str)

    _SKIP = {"", "nan", "none"}
    _PRIORITY = {"debit", "credit", "dr", "cr", "balance", "particulars", "account", "trial balance"}

    best_sheet = xl.sheet_names[0]
    best_score = -1

    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, dtype=str)
            if df.empty or len(df) < 3:
                continue

            # Score = number of non-empty cells + bonus if contains financial keywords
            all_text = " ".join(
                str(v).strip().lower()
                for v in df.values.flatten()
                if str(v).strip().lower() not in _SKIP
            )
            keyword_bonus = sum(20 for kw in _PRIORITY if kw in all_text)
            data_cells = sum(
                1 for v in df.values.flatten()
                if str(v).strip().lower() not in _SKIP
            )
            score = data_cells + keyword_bonus

            if score > best_score:
                best_score = score
                best_sheet = sheet
        except Exception:
            continue

    logger.info(f"Auto-selected sheet: '{best_sheet}' (score={best_score})")
    return pd.read_excel(file_path, sheet_name=best_sheet, header=None, dtype=str)


async def get_column_suggestions_with_llm(
    file_path: str,
    report_type: Optional[str],
    mapper_fields: list[str],
) -> dict:
    """
    Enhanced column suggestions: regex first, LLM fallback for unmapped columns.

    Returns:
        {
          "columns": [...],
          "suggestions": {standard_field: raw_col | None},   # regex results
          "llm_suggestions": {raw_col: system_field},        # LLM high-confidence
          "llm_questions": [{"column", "question", "options"}],
        }
    """
    from core.llm_manager import get_llm_provider

    basic = get_column_suggestions(file_path)
    columns: list[str] = basic["columns"]
    suggestions: dict = basic["suggestions"]

    if not mapper_fields or not columns:
        return {**basic, "llm_suggestions": {}, "llm_questions": []}

    # Columns already matched by regex (values of suggestions dict)
    regex_matched: set[str] = {v for v in suggestions.values() if v is not None}
    unmapped_cols = [c for c in columns if c not in regex_matched]

    if not unmapped_cols:
        return {**basic, "llm_suggestions": {}, "llm_questions": []}

    prompt = (
        f"You are a financial data expert. Raw column names from a trial balance file: {unmapped_cols}\n"
        f"Target system fields for a {report_type or 'financial'} report: {mapper_fields}\n\n"
        "For each raw column:\n"
        "- If you are >80% confident it maps to a system field, add it to 'mappings'.\n"
        "- If unsure, add a plain-English question (avoid finance jargon) to 'questions' with 2-4 simple options.\n"
        "- Ignore columns clearly irrelevant to financial data (e.g. dates, sequence numbers).\n\n"
        "Respond ONLY with valid JSON, no markdown:\n"
        '{"mappings": {"raw_col": "system_field"}, '
        '"questions": [{"column": "raw_col", "question": "...", "options": ["opt1", "opt2"]}]}'
    )

    try:
        llm = get_llm_provider(None)
        response = await llm.chat(
            [
                {"role": "system", "content": "You are a financial data mapping assistant. Respond with valid JSON only."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=1000,
        )
        raw = response.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        result = json.loads(raw.strip())
        llm_suggestions: dict[str, str] = {
            col: field
            for col, field in result.get("mappings", {}).items()
            if field in mapper_fields
        }
        llm_questions: list[dict] = [
            q for q in result.get("questions", [])
            if isinstance(q, dict) and "column" in q and "question" in q
        ]
        return {
            "columns": columns,
            "suggestions": suggestions,
            "llm_suggestions": llm_suggestions,
            "llm_questions": llm_questions,
        }
    except Exception as exc:
        logger.warning(f"LLM column mapping failed, returning regex-only results: {exc}")
        return {**basic, "llm_suggestions": {}, "llm_questions": []}


def get_column_suggestions(file_path: str) -> dict:
    """
    Return column headers + suggested mappings for the frontend mapper UI.
    Handles multi-sheet workbooks by auto-selecting the most data-rich sheet.

    Returns:
        {
          "columns": ["col1", "col2", ...],
          "suggestions": {"account_code": "col1", ...}   # may have None values
        }
    """
    import pandas as pd

    ext = file_path.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        df_raw = pd.read_csv(file_path, header=None, dtype=str, keep_default_na=False)
    else:
        df_raw = _best_sheet_df(file_path)

    header_row = _detect_header_row(df_raw)
    headers = [str(v).strip() for v in df_raw.iloc[header_row].tolist() if str(v).strip().lower() not in ("", "nan")]

    # If no headers detected (all empty), return generic labels based on column count
    if not headers:
        headers = [f"Column {i+1}" for i in range(min(df_raw.shape[1], 5))]

    suggestions = _fuzzy_map_columns(headers)
    return {"columns": headers, "suggestions": suggestions}
