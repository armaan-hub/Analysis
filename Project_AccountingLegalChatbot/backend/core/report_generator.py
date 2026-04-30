"""
Report Generator – Handles IFRS Financial Statements, UAE VAT Returns, and Corporate Tax Returns.
Generates output as Excel / PDF formats based on input data.
"""

import asyncio
import logging
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, borders, Border
from pathlib import Path
from config import settings

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Handles generating financial and tax reports."""
    
    def __init__(self):
        self.output_dir = Path(settings.upload_dir) / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_thick_border(self) -> dict:
        """Helper to return thick borders for Excel."""
        thick = borders.Side(style="medium", color="000000")
        return {"top": thick, "bottom": thick, "left": thick, "right": thick}

    async def generate_ifrs_statement(self, data: dict, company_name: str) -> str:
        """
        Generate IFRS Income Statement and Balance Sheet from trial balance data.
        Returns the path to the generated Excel file.
        """
        def _sync():
            wb = openpyxl.Workbook()

            # ── Income Statement ────────────────────────────────────
            ws_income = wb.active
            ws_income.title = "Income Statement"

            # Headers
            ws_income["A1"] = company_name
            ws_income["A1"].font = Font(bold=True, size=14)
            ws_income["A2"] = f"Statement of Comprehensive Income - {datetime.now().strftime('%Y')}"
            ws_income["A2"].font = Font(bold=True)

            row = 4
            ws_income.cell(row=row, column=1, value="Account").font = Font(bold=True)
            ws_income.cell(row=row, column=2, value="Amount (AED)").font = Font(bold=True)
            row += 1

            # Data (mock processing)
            total_revenue = data.get("revenue", 0)
            cogs = data.get("cogs", 0)
            gross_profit = total_revenue - cogs

            expenses = data.get("operating_expenses", 0)
            net_income = gross_profit - expenses

            records = [
                ("Revenue", total_revenue),
                ("Cost of Goods Sold", -cogs),
                ("Gross Profit", gross_profit),
                ("Operating Expenses", -expenses),
                ("Net Income", net_income)
            ]

            for name, amount in records:
                ws_income.cell(row=row, column=1, value=name)
                c2 = ws_income.cell(row=row, column=2, value=amount)
                if "Profit" in name or "Income" in name:
                    ws_income.cell(row=row, column=1).font = Font(bold=True)
                    c2.font = Font(bold=True)
                    c2.border = Border(top=borders.Side(style='thin'), bottom=borders.Side(style='double'))
                row += 1

            # ── Balance Sheet ───────────────────────────────────────
            ws_bs = wb.create_sheet(title="Balance Sheet")
            ws_bs["A1"] = company_name
            ws_bs["A1"].font = Font(bold=True, size=14)
            ws_bs["A2"] = f"Statement of Financial Position as of {datetime.now().strftime('%d %b %Y')}"
            ws_bs["A2"].font = Font(bold=True)

            row = 4
            assets = data.get("assets", 0)
            liabilities = data.get("liabilities", 0)
            equity = assets - liabilities  # Simplified Accounting Equation

            bs_records = [
                ("ASSETS", ""),
                ("Total Assets", assets),
                ("LIABILITIES & EQUITY", ""),
                ("Total Liabilities", liabilities),
                ("Total Equity", equity),
                ("Total Liabilities & Equity", liabilities + equity)
            ]

            for name, amount in bs_records:
                ws_bs.cell(row=row, column=1, value=name)
                if amount != "":
                    ws_bs.cell(row=row, column=2, value=amount)
                if "ASSETS" == name or "LIABILITIES" in name or "Total" in name:
                    ws_bs.cell(row=row, column=1).font = Font(bold=True)
                    ws_bs.cell(row=row, column=2).font = Font(bold=True)
                row += 1

            # Save file
            filename = f"IFRS_Statement_{company_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            logger.info(f"Generated IFRS report: {filepath}")
            return str(filepath)

        return await asyncio.to_thread(_sync)

    async def generate_vat_return(self, transactions: list[dict], company_name: str, trn: str) -> str:
        """
        Generate UAE FTA VAT 201 Return based on input transaction data.
        Returns the path to the generated Excel file.
        """
        def _sync():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VAT Return (VAT201)"

            ws["A1"] = "UAE FTA VAT Return Form (VAT201)"
            ws["A1"].font = Font(bold=True, size=16, color="003366")

            ws["A3"] = f"Taxable Person: {company_name}"
            ws["A4"] = f"TRN: {trn}"
            ws["A5"] = (
                "DISCLAIMER: This is a system-generated estimate only. "
                "Per-emirate amounts require accurate emirate tagging on each transaction. "
                "Review with a qualified UAE tax advisor before FTA submission."
            )
            ws["A5"].font = Font(italic=True, color="CC0000", size=9)

            # Table Headers
            row = 7
            headers = ["Box", "Description", "Amount (AED)", "VAT Amount (AED)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

            row += 1

            # DISCLAIMER: This output is an estimate only. Review with a qualified tax advisor
            # before submission to the FTA. Per-emirate amounts require accurate emirate tagging
            # on each transaction; untagged transactions are reported as "Other Emirates" (Box 1g).

            # Emirate box map (FTA VAT 201 Boxes 1a–1g)
            EMIRATE_BOX = {
                "abu_dhabi":          ("1a", "Standard rated supplies in Abu Dhabi"),
                "dubai":              ("1b", "Standard rated supplies in Dubai"),
                "sharjah":            ("1c", "Standard rated supplies in Sharjah"),
                "ajman":              ("1d", "Standard rated supplies in Ajman"),
                "umm_al_quwain":      ("1e", "Standard rated supplies in Umm Al Quwain"),
                "ras_al_khaimah":     ("1f", "Standard rated supplies in Ras Al Khaimah"),
                "fujairah":           ("1g", "Standard rated supplies in Fujairah"),
            }
            DEFAULT_BOX = ("1g", "Standard rated supplies — Other / Unspecified Emirates")

            standard_sales_by_emirate: dict = {}
            for t in transactions:
                if t.get('type') == 'sale' and t.get('tax_rate') == 0.05:
                    emirate_key = (t.get('emirate') or '').lower().replace(' ', '_')
                    box, label = EMIRATE_BOX.get(emirate_key, DEFAULT_BOX)
                    entry = standard_sales_by_emirate.setdefault(box, {'label': label, 'amount': 0.0})
                    entry['amount'] += t.get('amount', 0)

            standard_sales = sum(e['amount'] for e in standard_sales_by_emirate.values())
            standard_sales_vat = standard_sales * 0.05

            standard_purchases = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'purchase' and t.get('tax_rate') == 0.05)
            standard_purchases_vat = standard_purchases * 0.05

            net_vat_due = standard_sales_vat - standard_purchases_vat

            # Build per-emirate rows (1a–1g), only include boxes that have amounts
            emirate_rows = [
                (box, entry['label'], entry['amount'], round(entry['amount'] * 0.05, 2))
                for box, entry in sorted(standard_sales_by_emirate.items())
            ]
            # If no transactions had emirate data, fall back to a single unspecified row
            if not emirate_rows:
                emirate_rows = [("1g", "Standard rated supplies (emirate unspecified — see disclaimer)", 0, 0)]

            rows = [
                *emirate_rows,
                ("2", "Tax Refunds provided to Tourists", 0, 0),
                ("3", "Supplies subject to the reverse charge", 0, 0),
                ("4", "Zero rated supplies", 0, "-"),
                ("5", "Exempt supplies", 0, "-"),
                ("6", "Goods imported into the UAE", 0, 0),
                ("7", "Adjustments to goods imported into the UAE", 0, 0),
                ("8", "Standard rated expenses", standard_purchases, standard_purchases_vat),
                ("9", "Supplies subject to the reverse charge provision", 0, 0),
                ("12", "Net VAT Due", "-", net_vat_due),
            ]

            for box, desc, amount, vat in rows:
                ws.cell(row=row, column=1, value=box)
                ws.cell(row=row, column=2, value=desc)
                ws.cell(row=row, column=3, value=amount)
                c4 = ws.cell(row=row, column=4, value=vat)
                if box == "12":
                    for i in range(1, 5):
                        ws.cell(row=row, column=i).font = Font(bold=True)
                    c4.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            filename = f"VAT201_{trn}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            logger.info(f"Generated VAT return: {filepath}")
            return str(filepath)

        return await asyncio.to_thread(_sync)

    async def generate_corp_tax_return(self, data: dict, company_name: str) -> str:
        """
        Generate UAE Corporate Tax baseline file.
        """
        def _sync():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Corporate Tax Calculation"

            ws["A1"] = "UAE Corporate Tax Calculation"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = f"Taxable Entity: {company_name}"

            row = 4

            # Standard CT logic (simplified)
            accounting_profit = data.get("net_income", 0)
            disallowed_expenses = data.get("disallowed_expenses", 0)
            exempt_income = data.get("exempt_income", 0)

            taxable_income = accounting_profit + disallowed_expenses - exempt_income

            # CT rate: 0% up to 375k, 9% above
            threshold = 375000
            tax_amount = 0
            if taxable_income > threshold:
                tax_amount = (taxable_income - threshold) * 0.09

            rows = [
                ("Accounting Net Profit", accounting_profit),
                ("Add: Non-deductible expenses", disallowed_expenses),
                ("Less: Exempt Income", -exempt_income),
                ("Taxable Income", taxable_income),
                ("Exempt Threshold", threshold),
                ("Amount Subject to 9% Rate", max(0, taxable_income - threshold)),
                ("Corporate Tax Payable", tax_amount)
            ]

            for desc, val in rows:
                ws.cell(row=row, column=1, value=desc)
                c2 = ws.cell(row=row, column=2, value=val)
                if "Taxable Income" in desc or "Tax Payable" in desc:
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    c2.font = Font(bold=True)
                row += 1

            filename = f"CT_Return_{company_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            logger.info(f"Generated Corporate Tax return: {filepath}")
            return str(filepath)

        return await asyncio.to_thread(_sync)

report_generator = ReportGenerator()
